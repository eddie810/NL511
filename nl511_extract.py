"""
nl511_extract.py
Fetches road conditions and traffic events from the 511NL (511nl.ca) API,
exports to CSV, Excel, and KML.

Usage:
    python3 nl511_extract.py

Dependencies:
    pip3 install requests pandas openpyxl --break-system-packages
"""

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
from xml.dom import minidom
from datetime import datetime, timezone
import sys

# ── Configuration ────────────────────────────────────────────────────────────
API_KEY   = "002e66c43dd1491abe476df3e2bf034f"          # ← replace with your 511NL API key
BASE_URL  = "https://511nl.ca/api/v2"

OUT_CSV_CONDITIONS = "nl511_road_conditions.csv"
OUT_CSV_EVENTS     = "nl511_events.csv"
OUT_XLSX           = "nl511_data.xlsx"
OUT_KML            = "nl511_roads.kml"

# ── Colour map (KML AABBGGRR) ─────────────────────────────────────────────
# Colours match the official NL511 legend exactly.
# KML format is AABBGGRR (Alpha, Blue, Green, Red — reversed vs HTML).
# Order matters: more-specific substrings must come before broader ones.
CONDITION_COLOURS = {
    "closed":                   "ff0000ff",   # red     (#FF0000)
    "travel not recommended":   "ff00d7ff",   # yellow  (#FFD700)
    "partly covered":           "ffcc99cc",   # mauve   (#CC99CC) — before "covered"
    "covered":                  "ffffff00",   # cyan    (#00FFFF) — catches "covered snow packed" etc.
    "compact snow":             "ffffff00",   # cyan    (#00FFFF)
    "bare wet":                 "ff000000",   # black
    "bare dry":                 "ff000000",   # black
    "poor visibility":          "ffffffff",   # white   (legend shows dashed — KML approximation)
}
DEFAULT_COLOUR = "ff888888"              # grey (No Report)


# ── Polyline decoder ─────────────────────────────────────────────────────────
def decode_polyline(encoded: str) -> list:
    """Decode a Google Maps encoded polyline into a list of (lat, lng) tuples."""
    coords, index, lat, lng = [], 0, 0, 0
    while index < len(encoded):
        for is_lng in (False, True):
            result, shift = 0, 0
            while True:
                b = ord(encoded[index]) - 63
                index += 1
                result |= (b & 0x1F) << shift
                shift += 5
                if b < 0x20:
                    break
            delta = ~(result >> 1) if result & 1 else result >> 1
            if is_lng:
                lng += delta
            else:
                lat += delta
        coords.append((lat / 1e5, lng / 1e5))
    return coords


# ── Colour lookup ─────────────────────────────────────────────────────────────
def condition_colour(condition: str) -> str:
    """Return the KML colour for a given condition string."""
    if not condition:
        return DEFAULT_COLOUR
    lower = condition.lower()
    for key, colour in CONDITION_COLOURS.items():
        if key in lower:
            return colour
    return DEFAULT_COLOUR


# ── API fetch ─────────────────────────────────────────────────────────────────
def fetch(endpoint: str) -> list:
    """
    Fetch JSON from the 511NL API.
    Tries v2 first; falls back to v3 if needed.
    Returns a list of raw dicts (empty list on failure).
    """
    for version in ("v2", "v3"):
        url = f"https://511nl.ca/api/{version}/get/{endpoint}"
        try:
            resp = requests.get(url, params={"key": API_KEY}, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                if isinstance(data, list):
                    return data
                if isinstance(data, dict):
                    # some endpoints wrap in a key
                    for v in data.values():
                        if isinstance(v, list):
                            return v
            print(f"  [{version}] HTTP {resp.status_code} for {endpoint}")
        except Exception as exc:
            print(f"  [{version}] Error fetching {endpoint}: {exc}")
    return []


# ── Normalisation ─────────────────────────────────────────────────────────────
def normalise_condition(raw: dict) -> dict:
    """Map a raw winterroads API record to a clean flat dict."""
    def get(*keys):
        for k in keys:
            if k in raw and raw[k] not in (None, ""):
                return raw[k]
        return None

    return {
        "id":                    str(get("Id", "ID", "id") or ""),
        "roadway_name":          get("RoadwayName", "Roadway Name", "roadwayName", "roadway_name"),
        "location_description":  get("LocationDescription", "Location Description",
                                     "locationDescription", "location_description"),
        "from_measure":          get("FromMeasure", "From Measure", "fromMeasure"),
        "to_measure":            get("ToMeasure",   "To Measure",   "toMeasure"),
        "direction":             get("Direction",   "direction"),
        "road_condition":        get("Primary Condition", "PrimaryCondition",
                                     "Primary-Condition",  "primaryCondition",
                                     "RoadCondition",      "road_condition"),
        "secondary_conditions":  get("Secondary Conditions", "SecondaryConditions",
                                     "Secondary-Conditions", "secondaryConditions"),
        "visibility":            get("Visibility",  "visibility"),
        "last_updated":          get("LastUpdated", "Last Updated", "lastUpdated",
                                     "UpdatedAt",   "updatedAt"),
        "_encoded_polyline":     get("EncodedPolyline", "Encoded Polyline",
                                     "encodedPolyline", "encoded_polyline",
                                     "polyline", "Polyline"),
    }


def normalise_event(raw: dict) -> dict:
    """Map a raw events API record to a clean flat dict."""
    def get(*keys):
        for k in keys:
            if k in raw and raw[k] not in (None, ""):
                return raw[k]
        return None

    return {
        "id":                    str(get("Id", "ID", "id") or ""),
        "event_type":            get("EventType",   "Event Type",   "eventType",   "event_type"),
        "sub_type":              get("SubType",     "Sub Type",     "subType",     "sub_type"),
        "roadway_name":          get("RoadwayName", "Roadway Name", "roadwayName", "roadway_name"),
        "location_description":  get("LocationDescription", "Location Description",
                                     "locationDescription", "location_description"),
        "direction":             get("Direction",   "direction"),
        "severity":              get("Severity",    "severity"),
        "start_time":            get("StartTime",   "Start Time",   "startTime",   "start_time"),
        "end_time":              get("EndTime",     "End Time",     "endTime",     "end_time"),
        "description":           get("Description", "description"),
        "last_updated":          get("LastUpdated", "Last Updated", "lastUpdated"),
        "_encoded_polyline":     get("EncodedPolyline", "Encoded Polyline",
                                     "encodedPolyline", "encoded_polyline",
                                     "polyline", "Polyline"),
    }


# ── DataFrame helpers ─────────────────────────────────────────────────────────
def to_dataframe(records: list) -> tuple:
    """
    Convert a list of normalised records to a (df_main, df_polylines) tuple.
    Columns starting with '_' are stripped from df_main and placed in df_polylines
    (with the underscore prefix removed).
    """
    if not records:
        return pd.DataFrame(), pd.DataFrame()

    df = pd.DataFrame(records)

    poly_cols = [c for c in df.columns if c.startswith("_")]
    main_cols = [c for c in df.columns if not c.startswith("_")]

    df_main = df[main_cols].copy()
    if poly_cols:
        df_poly = df[["id"] + poly_cols].copy()
        df_poly.columns = ["id"] + [c.lstrip("_") for c in poly_cols]
    else:
        df_poly = pd.DataFrame(columns=["id"])

    return df_main, df_poly


# ── KML builder ───────────────────────────────────────────────────────────────
def build_kml(df_main: pd.DataFrame, df_poly: pd.DataFrame,
              folder_name: str, condition_col: str) -> ET.Element:
    """
    Build a KML <Document> element with colour-coded LineString placemarks.

    Uses a LEFT join so segments without polylines are reported but not silently
    dropped.  id columns are cast to str to prevent int/str type-mismatch joins.
    """
    doc = ET.Element("Document")

    # ── Styles ──────────────────────────────────────────────────────────────
    all_colours = dict(CONDITION_COLOURS)
    all_colours["__default__"] = DEFAULT_COLOUR

    for cond_key, colour in all_colours.items():
        style_id = cond_key.replace(" ", "_")
        style = ET.SubElement(doc, "Style", id=style_id)
        line  = ET.SubElement(style, "LineStyle")
        ET.SubElement(line, "color").text = colour
        ET.SubElement(line, "width").text = "4"

    # ── Merge ────────────────────────────────────────────────────────────────
    if df_main.empty:
        return doc

    dm = df_main.copy()
    dp = df_poly.copy()
    dm["id"] = dm["id"].astype(str)
    dp["id"] = dp["id"].astype(str)

    merged = dm.merge(dp, on="id", how="left")   # LEFT join — keep all segments

    # ── Folder ───────────────────────────────────────────────────────────────
    folder = ET.SubElement(doc, "Folder")
    ET.SubElement(folder, "name").text = folder_name

    skipped = []

    for _, row in merged.iterrows():
        encoded = row.get("encoded_polyline", "")
        if not encoded or str(encoded).strip() in ("", "nan", "None"):
            skipped.append(
                f"  id={row.get('id')} | {row.get('roadway_name', '')} "
                f"| {row.get('location_description', '')} "
                f"| {condition_col}={row.get(condition_col, '')}"
            )
            continue

        condition = str(row.get(condition_col, "") or "").lower()
        colour    = condition_colour(condition)

        # Match to a style id
        style_id = "__default__"
        for key in CONDITION_COLOURS:
            if key in condition:
                style_id = key.replace(" ", "_")
                break

        coords = decode_polyline(str(encoded))
        if not coords:
            skipped.append(f"  id={row.get('id')} — polyline decode failed")
            continue

        pm = ET.SubElement(folder, "Placemark")
        ET.SubElement(pm, "name").text = (
            str(row.get("roadway_name", "")) + " — " +
            str(row.get("location_description", ""))
        )
        ET.SubElement(pm, "description").text = (
            f"{condition_col}: {row.get(condition_col, '')}\n"
            f"Direction: {row.get('direction', '')}\n"
            f"Last updated: {row.get('last_updated', '')}"
        )
        ET.SubElement(pm, "styleUrl").text = f"#{style_id}"

        ls   = ET.SubElement(pm, "LineString")
        ET.SubElement(ls, "tessellate").text = "1"
        ET.SubElement(ls, "coordinates").text = " ".join(
            f"{lng},{lat},0" for lat, lng in coords
        )

    if skipped:
        print(f"\n⚠  {len(skipped)} segment(s) in '{folder_name}' skipped "
              f"(no polyline data):")
        for s in skipped:
            print(s)

    return doc


def export_kml(conditions: list, events: list, path: str) -> None:
    """Write a KML file with Road Conditions and Events folders."""
    df_cond_main, df_cond_poly = to_dataframe(conditions)
    df_evt_main,  df_evt_poly  = to_dataframe(events)

    root = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
    doc  = ET.SubElement(root, "Document")
    ET.SubElement(doc, "name").text = (
        f"NL511 Roads — {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    )

    # Road conditions folder
    cond_doc = build_kml(df_cond_main, df_cond_poly,
                         "Road Conditions", "road_condition")
    for child in list(cond_doc):
        doc.append(child)

    # Events folder
    evt_doc = build_kml(df_evt_main, df_evt_poly,
                        "Events & Incidents", "event_type")
    for child in list(evt_doc):
        doc.append(child)

    xml_str = minidom.parseString(
        ET.tostring(root, encoding="unicode")
    ).toprettyxml(indent="  ")

    with open(path, "w", encoding="utf-8") as f:
        f.write(xml_str)

    print(f"✓  KML saved → {path}")


# ── Excel export ──────────────────────────────────────────────────────────────
def _style_header(ws):
    """Apply header styling to row 1 of a worksheet."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def export(conditions: list, events: list) -> None:
    """Write CSV and Excel files for conditions and events."""

    df_cond_main, df_cond_poly = to_dataframe(conditions)
    df_evt_main,  df_evt_poly  = to_dataframe(events)

    # ── CSVs ──────────────────────────────────────────────────────────────
    if not df_cond_main.empty:
        df_cond_main.to_csv(OUT_CSV_CONDITIONS, index=False)
        print(f"✓  CSV saved → {OUT_CSV_CONDITIONS}  ({len(df_cond_main)} rows)")

    if not df_evt_main.empty:
        df_evt_main.to_csv(OUT_CSV_EVENTS, index=False)
        print(f"✓  CSV saved → {OUT_CSV_EVENTS}  ({len(df_evt_main)} rows)")

    # ── Excel ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    def add_sheet(name, df):
        if df.empty:
            return
        ws = wb.create_sheet(name)
        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        _style_header(ws)

    add_sheet("Road Conditions",   df_cond_main)
    add_sheet("Events & Incidents", df_evt_main)
    add_sheet("Road Polylines",    df_cond_poly)
    add_sheet("Event Polylines",   df_evt_poly)

    # Metadata sheet
    ws_meta = wb.create_sheet("Metadata")
    meta_rows = [
        ("Generated",    datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
        ("Source",       "https://511nl.ca"),
        ("Conditions",   len(conditions)),
        ("Events",       len(events)),
    ]
    for r, (k, v) in enumerate(meta_rows, 1):
        ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=r, column=2, value=str(v))

    wb.save(OUT_XLSX)
    print(f"✓  Excel saved → {OUT_XLSX}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("── Fetching NL511 data ──────────────────────────────────────────")

    print("Fetching winter road conditions …")
    raw_cond = fetch("winterroads")
    print(f"  Got {len(raw_cond)} raw condition records")

    print("Fetching events …")
    raw_evt  = fetch("events")
    print(f"  Got {len(raw_evt)} raw event records")

    if raw_cond:
        # Debug: show first record (excluding polyline noise)
        sample = {k: v for k, v in raw_cond[0].items()
                  if "polyline" not in k.lower() and "encoded" not in k.lower()}
        print(f"\nSample condition record keys: {list(sample.keys())}")

    conditions = [normalise_condition(r) for r in raw_cond]
    events     = [normalise_event(r)     for r in raw_evt]

    # ── Summaries ──────────────────────────────────────────────────────────
    if conditions:
        df_cond, _ = to_dataframe(conditions)
        if "road_condition" in df_cond.columns:
            print("\nRoad condition breakdown:")
            print(df_cond["road_condition"].value_counts().to_string())
        if "visibility" in df_cond.columns:
            print("\nVisibility breakdown:")
            print(df_cond["visibility"].value_counts().to_string())

    export(conditions, events)
    export_kml(conditions, events, OUT_KML)

    print("\n── Done ─────────────────────────────────────────────────────────")


if __name__ == "__main__":
    main()
